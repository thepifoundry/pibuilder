"""
Extract ground truth data from Excel files into JSON fixtures.
Run once when Excel files change: uv run python scripts/extract_fixtures.py
Output: tests/fixtures/mpsb_ground_truth.json, tests/fixtures/mpsb_grid_data.json,
        tests/fixtures/indargarh_daily_max.json, tests/fixtures/indargarh_ground_truth.json
"""

import json
import numpy as np
import openpyxl
from pathlib import Path

RICALCS = Path(__file__).parent.parent / "docs" / "ricalcs"
MPSB_PATH = RICALCS / "MPSB Rainfall Analysis_China Region.xlsx"
INDARGARH_PATH = RICALCS / "Indargarh daily rainfall from 1990-2021.xlsx"
FIXTURES = Path(__file__).parent.parent / "tests" / "fixtures"


def extract_mpsb_ground_truth(ws) -> list[dict]:
    """
    Summary sheet columns (0-indexed):
      0=lat, 1=lon,
      2-4=Pearson(rp,chi,ks), 5-7=LP3(rp,chi,ks),
      8-10=Gumbel(rp,ks,chi) NOTE ks/chi swapped,
      11-13=Normal(rp,ks,chi), 14-16=LogNorm(rp,ks,chi),
      17=applicable_100yr, 18=applicable_method,
      19=selected_final_100yr, 20=adopted_method
    """
    rows = list(ws.iter_rows(values_only=True))
    results = []
    for row in rows[7:]:
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue
        results.append({
            "lat": float(row[0]),
            "lon": float(row[1]),
            "pearson": {"rp": row[2], "chi": row[3], "ks": row[4]},
            "lp3":     {"rp": row[5], "chi": row[6], "ks": row[7]},
            "gumbel":  {"rp": row[8], "ks": row[9],  "chi": row[10]},
            "normal":  {"rp": row[11], "ks": row[12], "chi": row[13]},
            "lognorm": {"rp": row[14], "ks": row[15], "chi": row[16]},
            "selected_rp_100yr": row[19],
            "adopted_method": row[20],
        })
    return results


def extract_mpsb_grid_data(ws) -> dict:
    """
    Returns {lat_lon_key: [20 annual max values]} from 'log pearson (ks test)' sheet.
    Stored as {"lat": float, "lon": float, "values": [float, ...]} list.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    lats = all_rows[0][1:]
    lons = all_rows[1][1:]
    data_rows = all_rows[2:22]  # 20 years

    grid = []
    for col_idx in range(len(lats)):
        lat = lats[col_idx]
        lon = lons[col_idx]
        if lat is None or lon is None:
            continue
        values = [data_rows[yr][col_idx + 1] for yr in range(20)]
        if any(v is None for v in values):
            continue
        grid.append({
            "lat": float(lat),
            "lon": float(lon),
            "values": [float(v) for v in values],
        })
    return grid


def extract_indargarh_daily_max(ws) -> list[float]:
    rows = list(ws.iter_rows(values_only=True))
    values = []
    for row in rows[10:42]:
        if row[2] is not None and isinstance(row[2], (int, float)):
            values.append(float(row[2]))
    return values


def extract_indargarh_ground_truth(ws) -> dict:
    """Returns {duration_hr: {return_period: intensity_mm_hr}}."""
    rows = list(ws.iter_rows(values_only=True))
    return_periods = [2, 5, 10, 50, 100]
    result = {}
    for row in rows[52:57]:
        if row[12] is None:
            continue
        dur = int(row[12])
        intensities = {T: row[14 + i] for i, T in enumerate(return_periods)}
        result[str(dur)] = {str(T): v for T, v in intensities.items()}
    return result


def main():
    FIXTURES.mkdir(parents=True, exist_ok=True)

    print("Loading MPSB workbook...")
    mpsb_wb = openpyxl.load_workbook(MPSB_PATH, data_only=True)

    gt = extract_mpsb_ground_truth(mpsb_wb["summary"])
    (FIXTURES / "mpsb_ground_truth.json").write_text(json.dumps(gt, indent=2))
    print(f"  mpsb_ground_truth.json: {len(gt)} grid points")

    grid = extract_mpsb_grid_data(mpsb_wb["log pearson (ks test)"])
    (FIXTURES / "mpsb_grid_data.json").write_text(json.dumps(grid, indent=2))
    print(f"  mpsb_grid_data.json: {len(grid)} grid points")

    print("Loading Indargarh workbook...")
    ind_wb = openpyxl.load_workbook(INDARGARH_PATH, data_only=True)

    daily_max = extract_indargarh_daily_max(ind_wb["Rainfall Analysis"])
    (FIXTURES / "indargarh_daily_max.json").write_text(json.dumps(daily_max, indent=2))
    print(f"  indargarh_daily_max.json: {len(daily_max)} years")

    gt_ind = extract_indargarh_ground_truth(ind_wb["Rainfall Analysis"])
    (FIXTURES / "indargarh_ground_truth.json").write_text(json.dumps(gt_ind, indent=2))
    print(f"  indargarh_ground_truth.json: {len(gt_ind)} durations")

    print("Done.")


if __name__ == "__main__":
    main()
