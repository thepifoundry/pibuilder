"""
TDD: LP3 100yr return period must match MPSB Excel 'LP3 100yr RP' column exactly.
Tolerance: 0.01 mm (floating point round-trip from Excel).
"""

import numpy as np
import pytest
from pibuilder.rainfall.frequency import fit_lp3, _unbiased_skewness, _wilson_hilferty_kt


TOLERANCE = 0.01   # mm


class TestLP3Core:
    """Unit tests for the LP3 building blocks."""

    def test_unbiased_skewness_matches_excel(self, mpsb_grid_data, mpsb_ground_truth):
        """Excel Cs column uses n/((n-1)(n-2)) skewness — verify for first 10 points."""
        import openpyxl
        from pathlib import Path
        wb = openpyxl.load_workbook(
            Path(__file__).parent.parent / "docs/ricalcs/MPSB Rainfall Analysis_China Region.xlsx",
            data_only=True
        )
        ws = wb["log pearson (ks test)"]
        all_rows = list(ws.iter_rows(values_only=True))
        excel_cs_row = all_rows[25]   # row 26 = Cs

        lats = all_rows[0][1:]
        lons = all_rows[1][1:]

        for col_idx in range(min(10, len(lats))):
            lat, lon = lats[col_idx], lons[col_idx]
            if (lat, lon) not in mpsb_grid_data:
                continue
            data = np.log(mpsb_grid_data[(lat, lon)])
            excel_cs = excel_cs_row[col_idx + 1]
            my_cs = _unbiased_skewness(data)
            assert abs(my_cs - excel_cs) < 1e-6, (
                f"Cs mismatch at ({lat},{lon}): got {my_cs:.8f}, expected {excel_cs:.8f}"
            )

    def test_wilson_hilferty_matches_excel_kt(self, mpsb_grid_data, mpsb_ground_truth):
        """Excel Kt column must match exact WH cube-root formula to 6 decimal places."""
        import openpyxl
        from pathlib import Path
        wb = openpyxl.load_workbook(
            Path(__file__).parent.parent / "docs/ricalcs/MPSB Rainfall Analysis_China Region.xlsx",
            data_only=True
        )
        ws = wb["log pearson (ks test)"]
        all_rows = list(ws.iter_rows(values_only=True))
        excel_cs_row = all_rows[25]
        excel_kt_row = all_rows[27]
        lats = all_rows[0][1:]
        lons = all_rows[1][1:]

        for col_idx in range(min(20, len(lats))):
            lat, lon = lats[col_idx], lons[col_idx]
            excel_cs = excel_cs_row[col_idx + 1]
            excel_kt = excel_kt_row[col_idx + 1]
            if excel_cs is None or excel_kt is None:
                continue
            my_kt = _wilson_hilferty_kt(excel_cs, exceedance_prob=0.01)
            assert abs(my_kt - excel_kt) < 1e-6, (
                f"Kt mismatch at ({lat},{lon}): got {my_kt:.8f}, expected {excel_kt:.8f}"
            )


class TestLP3GridPoints:
    """LP3 100yr RP must match Excel for every MPSB grid point."""

    def test_lp3_first_grid_point(self, mpsb_grid_data, mpsb_ground_truth):
        """Regression: first grid point (27.875, 89.125) = 177.138 mm."""
        data = mpsb_grid_data[(27.875, 89.125)]
        rp, _ = fit_lp3(data)
        expected = 177.13777370520108
        assert abs(rp - expected) < TOLERANCE, f"LP3 RP: got {rp:.4f}, expected {expected:.4f}"

    def test_lp3_all_grid_points(self, mpsb_grid_data, mpsb_ground_truth):
        """LP3 100yr RP must match Excel LP3 column for all grid points."""
        failures = []
        for gt in mpsb_ground_truth:
            lat, lon = gt["lat"], gt["lon"]
            expected_lp3 = gt["lp3"]["rp"]
            if expected_lp3 is None or (lat, lon) not in mpsb_grid_data:
                continue
            data = mpsb_grid_data[(lat, lon)]
            rp, _ = fit_lp3(data)
            diff = abs(rp - expected_lp3)
            if diff > TOLERANCE:
                failures.append(f"({lat},{lon}): got {rp:.4f}, expected {expected_lp3:.4f}, diff={diff:.4f}")

        assert not failures, f"LP3 mismatch at {len(failures)} grid points:\n" + "\n".join(failures[:10])
